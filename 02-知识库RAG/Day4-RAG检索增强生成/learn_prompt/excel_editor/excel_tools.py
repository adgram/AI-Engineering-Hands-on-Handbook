import json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

CURRENT_FILE = None
CURRENT_DATA = None

def set_file(filepath):
    global CURRENT_FILE
    CURRENT_FILE = filepath

def read_table(filepath, sheet_name=None, max_rows=20):
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            rows.append([str(cell) if cell is not None else "" for cell in row])
            if i > max_rows:
                break
        wb.close()
        result = "\n".join([" | ".join(r) for r in rows])
        global CURRENT_DATA
        CURRENT_DATA = {"sheet": ws.title, "rows": len(list(ws.iter_rows())), "cols": ws.max_column}
        return result
    except Exception as e:
        return f"读取失败: {e}"

def write_cell(filepath, cell, value, sheet_name=None):
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws[cell] = value
        wb.save(filepath)
        wb.close()
        return json.dumps({"status": "ok", "cell": cell, "value": str(value)})
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)})

def write_formula(filepath, cell, formula, sheet_name=None):
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws[cell] = formula
        wb.save(filepath)
        wb.close()
        return json.dumps({"status": "ok", "cell": cell, "formula": formula})
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)})

def set_cell_style(filepath, cell, bold=False, font_color=None, fill_color=None, sheet_name=None):
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        cell_obj = ws[cell]
        if bold:
            cell_obj.font = Font(bold=True)
        if fill_color:
            cell_obj.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        wb.save(filepath)
        wb.close()
        return json.dumps({"status": "ok", "cell": cell})
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)})

def insert_row(filepath, row_idx, values, sheet_name=None):
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.insert_rows(row_idx)
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        wb.save(filepath)
        wb.close()
        return json.dumps({"status": "ok", "row": row_idx, "values": values})
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)})

def auto_sum(filepath, range_str, target_cell, sheet_name=None):
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws[target_cell] = f"=SUM({range_str})"
        wb.save(filepath)
        wb.close()
        return json.dumps({"status": "ok", "formula": f"=SUM({range_str})", "target": target_cell})
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)})

def merge_cells(filepath, range_str, value, sheet_name=None):
    try:
        wb = load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        ws.merge_cells(range_str)
        if value:
            top_left = range_str.split(":")[0]
            ws[top_left] = value
        wb.save(filepath)
        wb.close()
        return json.dumps({"status": "ok", "range": range_str, "value": value})
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)})

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_table",
            "description": "读取 Excel 表格内容（前 20 行），了解当前数据",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {"type": "string", "description": "Excel 文件路径"},
                    "sheet_name": {"type": "string", "description": "工作表名（可选）"}
                },
                "required": ["filepath"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_cell",
            "description": "向单元格写入值（文本/数字）",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {"type": "string", "description": "Excel 文件路径"},
                    "cell": {"type": "string", "description": "单元格地址，如 A1, B2"},
                    "value": {"type": "string", "description": "要写入的值"},
                    "sheet_name": {"type": "string", "description": "工作表名（可选）"}
                },
                "required": ["filepath", "cell", "value"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_formula",
            "description": "向单元格写入 Excel 公式",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {"type": "string", "description": "Excel 文件路径"},
                    "cell": {"type": "string", "description": "单元格地址"},
                    "formula": {"type": "string", "description": "Excel 公式，如 =A1+B1"},
                    "sheet_name": {"type": "string", "description": "工作表名（可选）"}
                },
                "required": ["filepath", "cell", "formula"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "set_cell_style",
            "description": "设置单元格样式（加粗、字体颜色、填充色）",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {"type": "string", "description": "Excel 文件路径"},
                    "cell": {"type": "string", "description": "单元格地址"},
                    "bold": {"type": "boolean", "description": "是否加粗"},
                    "font_color": {"type": "string", "description": "字体颜色，如 FFFFFF"},
                    "fill_color": {"type": "string", "description": "填充色，如 4472C4"},
                    "sheet_name": {"type": "string", "description": "工作表名（可选）"}
                },
                "required": ["filepath", "cell"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "insert_row",
            "description": "在指定位置插入一行数据",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {"type": "string", "description": "Excel 文件路径"},
                    "row_idx": {"type": "integer", "description": "行号（从 1 开始）"},
                    "values": {"type": "array", "items": {"type": "string"}, "description": "每列的值列表"},
                    "sheet_name": {"type": "string", "description": "工作表名（可选）"}
                },
                "required": ["filepath", "row_idx", "values"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "auto_sum",
            "description": "对指定区域求和，结果写入目标单元格",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {"type": "string", "description": "Excel 文件路径"},
                    "range_str": {"type": "string", "description": "求和区域，如 B2:B10"},
                    "target_cell": {"type": "string", "description": "结果单元格地址"},
                    "sheet_name": {"type": "string", "description": "工作表名（可选）"}
                },
                "required": ["filepath", "range_str", "target_cell"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "merge_cells",
            "description": "合并单元格，并在左上角单元格写入值",
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {"type": "string", "description": "Excel 文件路径"},
                    "range_str": {"type": "string", "description": "合并区域，如 A1:C1"},
                    "value": {"type": "string", "description": "合并后左上角单元格的值"},
                    "sheet_name": {"type": "string", "description": "工作表名（可选）"}
                },
                "required": ["filepath", "range_str"]
            }
        }
    }
]

FUNCTIONS = {
    "read_table": lambda **k: read_table(**k),
    "write_cell": lambda **k: write_cell(**k),
    "write_formula": lambda **k: write_formula(**k),
    "set_cell_style": lambda **k: set_cell_style(**k),
    "insert_row": lambda **k: insert_row(**k),
    "auto_sum": lambda **k: auto_sum(**k),
    "merge_cells": lambda **k: merge_cells(**k),
}
