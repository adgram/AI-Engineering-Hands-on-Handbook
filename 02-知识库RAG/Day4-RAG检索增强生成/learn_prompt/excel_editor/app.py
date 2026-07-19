import os, json, tempfile
from pathlib import Path
import gradio as gr
from openpyxl import Workbook
from llm_client import LLMClient
from prompts import SYSTEM_PROMPT
from excel_tools import TOOLS, FUNCTIONS, set_file

client = LLMClient()

def create_demo_file():
    demo_path = os.path.join(tempfile.gettempdir(), "demo_sales.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "销售数据"
    headers = ["月份", "产品", "销量", "单价", "销售额"]
    ws.append(headers)
    data = [
        ["1月", "会员卡", 120, 299],
        ["1月", "体验课", 45, 199],
        ["2月", "会员卡", 98, 299],
        ["2月", "体验课", 67, 199],
        ["3月", "会员卡", 150, 299],
        ["3月", "体验课", 88, 199],
    ]
    for row in data:
        ws.append(row)
    wb.save(demo_path)
    return demo_path

def read_excel_preview(filepath):
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append([cell if cell is not None else "" for cell in row])
    wb.close()
    return headers, data

def execute_tool(name, args):
    handler = FUNCTIONS.get(name)
    if handler:
        return handler(**args)
    return json.dumps({"error": f"未知工具: {name}"})

def process_message(filepath, user_input):
    set_file(filepath)
    if not os.path.exists(filepath):
        return "文件不存在", filepath

    from excel_tools import read_table
    table_context = read_table(filepath)

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": f"当前表格内容：\n{table_context}\n\n用户需求：{user_input}"}
    ]

    max_rounds = 8
    final_reply = ""
    for round_num in range(max_rounds):
        response = client.chat(messages=messages, tools=TOOLS, tool_choice="auto")
        message = response.choices[0].message

        if not message.tool_calls:
            final_reply = message.content
            break

        messages.append(message)
        for tc in message.tool_calls:
            name = tc.function.name
            args = json.loads(tc.function.arguments)
            args["filepath"] = filepath
            result = execute_tool(name, args)
            messages.append({"role": "tool", "tool_call_id": tc.id, "content": result})

    if not final_reply:
        final_reply = "操作完成（已达最大轮数）"

    headers, data = read_excel_preview(filepath)
    return final_reply, filepath, gr.Dataframe(value=data, headers=headers)

def upload_and_process(file, user_input):
    if file is None:
        filepath = create_demo_file()
        action = "已加载示例销售表"
    else:
        filepath = file.name
        action = f"已上传: {Path(file.name).name}"

    if not user_input.strip():
        headers, data = read_excel_preview(filepath)
        return action, filepath, gr.Dataframe(value=data, headers=headers)

    result, fp, table = process_message(filepath, user_input)
    return f"✓ {result}", fp, table

def create_file_from_template():
    fp = create_demo_file()
    headers, data = read_excel_preview(fp)
    msg = "已创建示例销售表（1-3 月会员卡和体验课销售数据）"
    return msg, fp, gr.Dataframe(value=data, headers=headers)

CSS = """
    .file-section { min-height: 200px; }
    .action-btn { min-width: 100%; }
    .result-box { border-left: 4px solid #4CAF50; padding-left: 10px; }
"""

with gr.Blocks(title="AI 表格编辑助手") as demo:
    gr.Markdown("""
    # 📊 AI 表格编辑助手
    用自然语言操作 Excel 表格，无需记忆公式和菜单路径。
    """)

    file_state = gr.State("")

    with gr.Row(equal_height=True):
        with gr.Column(scale=1, min_width=280):
            gr.Markdown("### 📁 文件")
            file_input = gr.File(label="上传 .xlsx 文件", file_types=[".xlsx"],
                                 height=160, elem_classes="file-section")
            create_btn = gr.Button("📄 创建示例表格", variant="secondary",
                                   elem_classes="action-btn")

        with gr.Column(scale=2, min_width=480):
            gr.Markdown("### 💬 说你想做什么")
            gr.Examples(
                examples=[
                    ["在 E 列写入公式计算销售额=销量×单价，表头写'销售额'"],
                    ["把标题行加粗，背景改成蓝色"],
                    ["在表格最下面加一行汇总，用公式求和"],
                    ["在 3 月数据后面插入一行：3月/团购课/200/99"],
                    ["做个全年汇总行，用深灰色背景"],
                ],
                inputs=[],
                label="💡 试试这些指令",
            )
            with gr.Row():
                user_input = gr.Textbox(
                    label="输入指令",
                    placeholder="例如：在E列插入公式 销售额=销量×单价",
                    lines=2,
                    scale=4,
                )
                send_btn = gr.Button("🚀 执行", variant="primary",
                                     scale=1, min_width=100)

    result_text = gr.Textbox(label="操作结果", lines=2, interactive=False,
                             elem_classes="result-box")

    table_preview = gr.Dataframe(label="表格预览", interactive=False,
                                 wrap=True)

    gr.Markdown("### 使用说明")
    gr.Markdown("""
    - **上传 Excel** 或点击「创建示例表格」开始
    - **输入自然语言指令**，AI 会自动拆解任务并操作表格
    - **示例指令**：计算销售额、加汇总行、设置格式、插入数据
    - 每次操作后表格预览会自动刷新
    """)

    send_btn.click(
        fn=upload_and_process,
        inputs=[file_input, user_input],
        outputs=[result_text, file_state, table_preview]
    )

    create_btn.click(
        fn=create_file_from_template,
        inputs=[],
        outputs=[result_text, file_state, table_preview]
    )

if __name__ == "__main__":
    demo.launch(theme=gr.themes.Soft(), css=CSS)
